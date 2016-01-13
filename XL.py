"""
    Written By Marcus Whelan for Senior design 1 project 2
    This file works with the Excel documents for the classes
"""
import openpyxl
import sys, csv,os,os.path
from openpyxl import *
from openpyxl.worksheet import Worksheet
from time import localtime,strftime
from sys import argv
import RDC
import mail

# ----------------Make data excel sheet -------------------
def classData():
    wb = Workbook()
    data_sheet = wb['Sheet']
    data_sheet['A1'] = "ClassN"
    wb.save("thisClass.xlsx")
    return
"""
=======================================================
                Constructors listed first
=======================================================
"""
# set last student written in kyboard
def setLstudent(Name):
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb['Sheet']
    sheet0['D1'].value = Name
    wb.save("thisClass.xlsx")
    return
# get last student written in keyboard
def getLstudent():
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb['Sheet']
    name = sheet0['D1'].value
    return name
# set last class written in keyboard
def setLclass(CName):
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb['Sheet']
    sheet0['D2'].value = CName
    wb.save("thisClass.xlsx")
    return
# get last class written in keyboard
def getLclass():
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb['Sheet']
    name = sheet0['D2'].value
    return name
# get list of all Students in class
def getStudents():
    array = []
    #load class data sheet to get class name
    ClassN = getClass()
    # load the students in to a list to return
    ws = load_workbook(ClassN+".xlsx")
    sheet = ws["Sheet1"]
    totalstn = getTstudents()
    w = 5
    while(True):
        temp = sheet.cell(row=w,column = 1)
        if(temp.value is not None):
            array.append(temp.value)
            w = w+1;
        elif(w >= totalstn+5):
            break;
        w = w+1;
    return array
# get list of all classes
def getClasses():
    array = []
    # load class data sheet.
    wb = load_workbook("thisClass.xlsx")
    sheet0 =wb['Sheet']
    w = 1
    while(True):
        temp= sheet0.cell(row= w,column = 2)
        if(temp.value is not None):
            array.append(temp.value)
        elif(w >= 20):
            break;
        w=w+1
    return array
# set current class
def setClass(ClassN):
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb["Sheet"]
    sheet0['A1'].value = ClassN
    wb.save("thisClass.xlsx")
    return
# get current class
def getClass():
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb["Sheet"]
    this_Class = sheet0['A1'].value
    return this_Class
# get Number of classes
def getNclasses():
    wb = load_workbook("thisClass.xlsx")
    sheet = wb['Sheet']
    classn = sheet.cell(row=5,column=1).value
    return classn
# get total number of students after all added and none deleted
def getTstudents():
    ClassN = getClass()
    wb=load_workbook(ClassN+".xlsx")
    sheet = wb['Sheet']
    totalstn = sheet.cell(row=7,column=1).value
    return int(totalstn)
# get number of students still in class/ after dropouts
def getNstudents():
    ClassN = getClass()
    wb= load_workbook(ClassN+".xlsx")
    sheet = wb['Sheet']
    stn = sheet.cell(row=5,column=1).value
    return int(stn)
# get hours in class
def getH():
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet1 = wb['Sheet1']
    temp = sheet1.cell(row=1,column=3).value
    return str(temp)
# get minutes in class
def getM():
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet1 = wb['Sheet1']
    temp = sheet1.cell(row=1,column=4).value
    return str(temp)
# change Hour time for class
def changeH(H):
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet0 = wb['Sheet']
    lectures = sheet0['A2'].value
    for x in range(1,(lectures+1)):
        sheet = wb["Sheet"+str(x)]
        temp = sheet.cell(row=1,column=3)
        temp.value = int(H)
    wb.save(ClassN+".xlsx")
    return
# change Minutes in class
def changeM(M):
    ClassN = getClass() 
    wb = load_workbook(ClassN+".xlsx")
    sheet0 = wb['Sheet']
    lectures = sheet0['A2'].value
    lectures = int(lectures)
    for x in range(1,(lectures+1)):
        sheet = wb["Sheet"+str(x)]
        temp = sheet.cell(row=1,column=4)
        temp.value = int(M)
    wb.save(ClassN+".xlsx")
    return
# set Last page visited
def setLastPage(last):
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb["Sheet"]
    sheet0['A4'].value = last
    wb.save("thisClass.xlsx")
    return
# get last page visited
def getPage():
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb["Sheet"]
    Page = sheet0['A4'].value
    return Page
# Set the lecture day
def setLectDay(number):
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet = wb['Sheet']
    sheet['A9'].value = number
    wb.save(ClassN+".xlsx")
    return
# increment the lecture day
def incLectDay():
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet= wb['Sheet']
    day = sheet['A9'].value
    sheet['A9'].value = day+1
    wb.save(ClassN+".xlsx")
    return
# get the lecture day after increments 
def getLectDay():
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet = wb['Sheet']
    day = sheet['A9'].value
    return int(day)
# get the lecture day for the actual class
def getLD():
    ClassN=getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet = wb['Sheet']
    lecture = sheet['A2'].value
    return int(lecture)
# sets class start time for that lecture
def setClassS():
    classStart = RDC.getCT()
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    day = getLectDay()
    sheetday = wb["Sheet"+str(day)]
    sheetday.cell(row=1,column=2).value = classStart
    wb.save(ClassN+".xlsx")
    return
# gets the tardy time set by professor or creation
def getTTime():
    ClassN= getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet = wb['Sheet']
    time = sheet.cell(row=11,column=1).value
    return int(time)
# sets the Tardy time for class
def setTTime(time):
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet = wb['Sheet']
    sheet.cell(row=11,column=1).value = time
    wb.save(ClassN+".xlsx")
    return
# gets the path of the current class to email out
def getClassPath():
    ClassN = getClass()
    path = ("/home/pi/Senior1/"+ClassN+".xlsx")
    return path
# Sets the users email when entered in
def setEmail(email):
    wb = load_workbook("thisClass.xlsx")
    sheet = wb['Sheet']
    sheet['E1'].value = email
    wb.save("thisClass.xlsx")
    return
# returns the current email to email the files to.
def getEmail():
    wb = load_workbook("thisClass.xlsx")
    sheet = wb['Sheet']
    email = sheet['E1'].value
    return email
# saves the new password
def setP(password):
    wb = load_workbook("thisClass.xlsx")
    sheet = wb['Sheet']
    sheet.cell(row=7,column=1).value = password
    wb.save("thisClass.xlsx")
    return
# get the current password
def getP():
    wb = load_workbook("thisClass.xlsx")
    sheet = wb['Sheet']
    password = sheet.cell(row=7,column=1).value
    return password
"""
=======================================================
            None Constructors
=======================================================
"""
# ----------------Make Excel sheet for class---------------
def createClass(ClassN,lectures):
    lectures = int(lectures)
    #first save new class in Class Data file
    ws = load_workbook("thisClass.xlsx")
    sheet0 = ws['Sheet']
    w=1
    while(True):
        temp = sheet0.cell(row=w,column = 2)
        if(temp.value is None):
            temp.value = ClassN
            break;
        else:
            w = w+1
    temp2 = sheet0.cell(row=1,column=1)
    if(temp2.value is None):
        temp2.value = ClassN
    #then add it to the number of classes there are
    classn = sheet0.cell(row=5,column=1)
    if(classn.value is None):
        classn.value = 1
    else:
        classn.value = classn.value +1
    ws.save("thisClass.xlsx")
    #Make class and save it
    if(os.path.isfile(ClassN+".xlsx")):
        f = open(ClassN+".xlsx","r+")
        f.close()
    else:
        wb = Workbook()
        class_lectrue = wb.active
        class_lecture = wb['Sheet']
        class_lecture['A1'] = "Lectures"
        class_lecture['B1'] = "Students"
        class_lecture['C1'] = "UID"
        class_lecture['A4'] = "Students N"
        class_lecture['A5'] = 1
        class_lecture['A6'] = "Past Total"
        class_lecture['A7'] = 1
        class_lecture.column_dimensions['B'].width = 30.0
        class_lecture['A2'] = lectures
        class_lecture['A8'] = "Today lect"
        class_lecture['A9'] = 1
        class_lecture['A10'] = "Tardy After"
        class_lecture['A11'] = 10
        for x in range(1,(lectures+1)):
            class_lecture = wb.create_sheet(title=("Sheet"+str(x)))
            class_lecture['A1'] = ClassN
            class_lecture['C1'] = 1
            class_lecture['D1'] = 15
            class_lecture['A4'] = "Student Name"
            class_lecture['B2'] = "classtime"
            class_lecture['C2'] = "hours"
            class_lecture['D2'] = "minutes"
            class_lecture['B4'] = "Photo"
            class_lecture['C4'] = "RFID"
            class_lecture['D4'] = "Time In"
            class_lecture['E4'] = "Tardy?"
            class_lecture.column_dimensions['A'].width = 30.0
        class_lecture = wb.create_sheet(title=("Sheet"+str(lectures+1)))
        class_lecture['A1'] = ClassN
        class_lecture['A3'] = "Student"
        class_lecture['B3'] = "Totals"
        class_lecture.column_dimensions['A'].width = 30.0
        wb.save(ClassN+".xlsx")
    return
#-----------------add UID for student------------------------
def addUID(UID):
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet0 = wb['Sheet']
    rw = 2
    while(True):
        name = sheet0.cell(row=rw,column=2)
        if(name.value == getLstudent()):
            uidcell = sheet0.cell(row=rw,column=3)
            uidcell.value = UID
            break;
        else:
            rw = rw +1
    wb.save(ClassN+".xlsx")
#------------------Add Student to Class ------------------------
def addStudent(Name):
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet0 = wb["Sheet"]
    lectures = sheet0['A2'].value
    lectures = int(lectures)
    stn = sheet0['A5'].value
    newstn = stn+1
    sheet0['A5'].value = newstn
    totalstn = sheet0['A7'].value
    newt = totalstn+1
    sheet0['A7'].value = newt
    #input student on first page for UID reference
    com = 1
    while(True):
        names = sheet0.cell(row=com,column=2)
        if(names.value is None):
            names.value = Name
            break;
        else:
            com = com +1
    #input the students name into each lecture
    for x in range(1,(lectures+1)):
        input_student_in_lecture = wb["Sheet"+str(x)]
        w = 5
        while(True):
            temp = input_student_in_lecture.cell(row=w,column = 1)
            if(temp.value is None):
                temp.value = Name
                break;
            else:
                w = w+1
    #input student in final sheet for totals
    sheet = wb['Sheet'+str(lectures+1)]
    last = 4
    while(True):
        line = sheet.cell(row=last,column=1)
        if(line.value is None):
            line.value = Name
            break;
        else:
            last = last+1
    wb.save(ClassN+".xlsx")
    return
#----------------Deletes student from class--------------------
def delStudent(Name):
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet0 = wb['Sheet']
    lectures = sheet0['A2'].value
    lectures = int(lectures)
    stn = sheet0['A5'].value
    newstn = stn-1
    sheet0['A5'].value = newstn
    #del student in first sheet
    ro = 2
    while(True):
        names = sheet0.cell(row=ro,column=2)
        if(names.value == Name):
            names.value = None
            while(True):
                r3w = sheet0.cell(row=ro,column=3)
                if(r3w.value is not None):
                    r3w.value = None
                    break;
                else:
                    break;
            break;
        else:
            ro = ro +1
    #delete the students name in each lecture
    for x in range(1,(lectures+1)):
        input_student_in_lecture = wb["Sheet"+str(x)]
        w = 5
        while(True):
            temp = input_student_in_lecture.cell(row=w,column = 1)
            if(temp.value == Name):
                temp.value = None
                y = 2
                ## maybe this will work??? delete everything in row
                while(True):
                    lisst = input_student_in_lecture.cell(row=w,column=y)
                    if(lisst.value is not None):
                        lisst.value = None
                        if(y == 6):
                            break;
                        y =y+1
                    else:
                        break;
                ##----------------
                break;
            else:
                w = w+1
    page = wb['Sheet'+str(lectures+1)]
    m = 4
    while(True):
        nam = page.cell(row=m,column=1)
        if(nam.value == Name):
            nam.value = None
            break;
        else:
            m= m+1
    wb.save(ClassN+".xlsx")
    return
#------------------Deletes class----------------------------------
def delClass(ClassN):
    array = []
    wb = load_workbook("thisClass.xlsx")
    sheet0 = wb['Sheet']
    classn = sheet0.cell(row=5,column=1)
    classn.value = classn.value - 1
    sec = sheet0.cell(row=1,column=1)
    if(sec.value == ClassN):
        sec.value = None
        array = getClasses()
        newTopClass = array[1]
        sec.value = newTopClass
    w = 1
    while(True):
        temp = sheet0.cell(row=w,column = 2)
        if(temp.value == ClassN):
            temp.value = None
            break;
        else:
            w = w+1
    wb.save("thisClass.xlsx")
    os.remove(ClassN+".xlsx")
    return
#----------------------Change class Name--------------------------
def changeCName(ClassN,OldN):
    #change the name in all the sheets of the class xl file
    os.rename(OldN+".xlsx",ClassN+".xlsx")
    wb = load_workbook(ClassN+".xlsx")
    sheet0 = wb["Sheet"]
    lectures = sheet0['A2'].value
    lectures = int(lectures)
    for x in range(1,(lectures+2)):
        sheet = wb["Sheet"+str(x)]
        temp = sheet.cell(row=1,column=1)
        temp.value = ClassN
    wb.save(ClassN+".xlsx")
    ws = load_workbook("thisClass.xlsx")
    sheet_0 = ws['Sheet']
    #now change the name in the data page
    w = 1
    while(True):
        temp = sheet_0.cell(row=w,column = 2)
        if(temp.value == OldN):
            temp.value = ClassN
            break;
        else:
            w = w+1
    name = sheet_0.cell(row=1,column=1)
    if(name.value == OldN):
        name.value = ClassN
        ws.save("thisClass.xlsx")
    return
#---------get Student from RFID number ------------
def getSfromUID(UID):
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheet0 = wb['Sheet']
    r = 2
    while(True):
        uids = sheet0.cell(row=r,column=3)
        if(uids.value == UID):
            name = sheet0.cell(row=r,column=2).value
            break;
        else:
            r = r+1
    return name
#------------Place student name and RFID in lecture day ----------
def placeStudent(UID,CT):
    day = getLectDay()
    isname = getSfromUID(UID)
    ClassN = getClass()
    wb = load_workbook(ClassN+".xlsx")
    sheetday = wb["Sheet"+str(day)]
    late = RDC.setTM(getTTime())
    starttime = CT
    time = RDC.getCT()
    c = 1
    r = 5
    while(True):
        student = sheetday.cell(row=r,column=1)
        if(student.value == isname):
            sheetday.cell(row=r,column=3).value = UID
            sheetday.cell(row=r,column=4).value = time
            if((time-starttime)> late):
                sheetday.cell(row=r,column=5).value = 0
            else:
                sheetday.cell(row=r,column=5).value = 1
            break;
        else:
            r = r +1
    wb.save(ClassN+".xlsx")
    return
#-----------Send the email out for the current class----------------
def emailCXL():
    path = getClassPath()
    email = getEmail()
    mail.sendXL(path,email)
    return
    



